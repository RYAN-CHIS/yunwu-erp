import psycopg2
from openpyxl import load_workbook

DATABASE_URL = 'postgresql://neondb_owner:npg_cAas8kuHmrO0@ep-polished-unit-ajk5rq34.c-3.us-east-2.aws.neon.tech/neondb?sslmode=require'

wb = load_workbook('/Users/ryan/Desktop/进货清单_新版.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]
rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0]]
print(f"Excel 读取：{len(rows)} 条数据")

conn = psycopg2.connect(DATABASE_URL)
conn.autocommit = True
cur = conn.cursor()

# 按外键依赖顺序清空
for table in ['purchase_records', 'bom', 'inventory_transactions', 'raw_materials']:
    try:
        cur.execute(f'DELETE FROM {table}')
        print(f"  清空 {table}：{cur.rowcount} 条")
    except Exception as e:
        print(f"  清空 {table} 失败：{e}")

# 形状映射（Excel 中文 → 中文，与系统一致）
SHAPE_MAP = {
    '圆珠': '圆珠', '挂饰': '挂饰', '三通': '三通',
    '单面珠': '单面珠', '隔片': '隔片', '老型': '老型',
    '汉堡珠': '汉堡珠', '方糖': '方糖', '桶珠': '桶珠',
    '随形': '随形', '方块': '方块',
}

# 中文品类 → MaterialType 枚举值
CATEGORY_TO_ENUM = {
    '水晶': 'BEAD',
    '沉香': 'INCENSE', '老山檀': 'INCENSE',
    '蜜蜡': 'BEAD', '南红': 'BEAD', '猛犸': 'BEAD',
    '青金石': 'BEAD', '堇青石': 'BEAD',
    '大漆珠': 'CERAMIC', '皮绳': 'CORD', '编绳': 'CORD',
    '花丝云锦带': 'CORD', '陶瓷': 'CERAMIC', '金属': 'METAL',
    '隔片': 'CERAMIC', '吊坠': 'CERAMIC',
}

total_cost_sum = 0.0
success = 0
failures = []

for i, row in enumerate(rows):
    try:
        code          = str(row[0]).strip()
        category_cn   = str(row[1]).strip() if row[1] else ''
        name          = str(row[2]).strip() if row[2] else ''
        supplier      = str(row[3]).strip() if row[3] else None
        spec          = str(int(row[4])) if row[4] and str(row[4]).strip() else None
        shape_raw     = str(row[5]).strip() if row[5] else ''
        remark        = str(row[6]).strip() if row[6] and str(row[6]).strip() else None
        mode          = str(row[7]).strip() if row[7] else ''
        qty           = float(row[8]) if row[8] else None
        price         = float(row[9]) if row[9] else None
        unit_cell     = str(row[10]).strip() if row[10] else ''
        total_price   = float(row[11]) if row[11] and isinstance(row[11], (int, float)) else None
        beads_val     = float(row[12]) if row[12] and isinstance(row[12], (int, float)) else None
        weight_val    = float(row[13]) if row[13] and isinstance(row[13], (int, float)) else None
        total_beads   = float(row[14]) if row[14] and isinstance(row[14], (int, float)) else None
        total_weight  = float(row[15]) if row[15] and isinstance(row[15], (int, float)) else None
        unit_cost     = float(row[16]) if row[16] and isinstance(row[16], (int, float)) else None

        if not code or not name:
            failures.append((i + 4, code, '缺少编码或名称'))
            continue

        # 形状
        shape = SHAPE_MAP.get(shape_raw, shape_raw) if shape_raw else None

        # materialType 枚举值
        mat_enum = CATEGORY_TO_ENUM.get(category_cn, 'OTHER')

        # 计价方式逻辑
        if mode == '按克':
            inv_unit = '克'
            purchase_unit = '克'
            conversion_rate = 1.0
            beads_per_strand = None
            weight_per_strand = weight_val or ((total_weight / qty) if (total_weight and qty and qty > 0) else None)
            remaining = total_weight if total_weight else 0
        elif mode == '按个':
            inv_unit = '个'
            purchase_unit = '个'
            conversion_rate = 1.0
            beads_per_strand = None
            weight_per_strand = None
            remaining = qty if qty else 0
            if total_price and qty and qty > 0:
                unit_cost = total_price / qty
        elif mode == '按串':
            inv_unit = '颗'
            purchase_unit = '串'
            conversion_rate = beads_val if beads_val else 1.0
            beads_per_strand = beads_val if beads_val else None
            weight_per_strand = None
            remaining = total_beads if total_beads else 0
        else:
            inv_unit = '个'
            purchase_unit = '个'
            conversion_rate = 1.0
            beads_per_strand = None
            weight_per_strand = None
            remaining = 0

        if not unit_cost or unit_cost <= 0:
            unit_cost = 0.01

        cur.execute("""
            INSERT INTO raw_materials
            (code, "materialType", category, name, supplier,
             "specification", shape, remark,
             "unitCost", "inventoryUnit", default_purchase_unit, default_conversion_rate,
             beads_per_strand, weight_per_strand, remaining, "status",
             created_at, updated_at)
            VALUES (%s,%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s, NOW(),NOW())
            RETURNING id
        """, (
            code,
            mat_enum,
            category_cn,
            name,
            supplier,
            spec,
            shape,
            remark,
            unit_cost,
            inv_unit,
            purchase_unit,
            conversion_rate,
            beads_per_strand,
            weight_per_strand,
            remaining,
            'READY',
        ))
        material_id = cur.fetchone()[0]

        # 写入入库流水
        if remaining and remaining > 0:
            cur.execute("""
                INSERT INTO inventory_transactions
                ("materialId", type, quantity, "beforeQty", "afterQty", remark, created_at)
                VALUES (%s, 'IN', %s, 0, %s, 'Excel导入', NOW())
            """, (material_id, remaining, remaining))

        if total_price:
            total_cost_sum += total_price
        success += 1

    except Exception as e:
        failures.append((i + 4, code if 'code' in dir() else str(row[0]), str(e)))

conn.close()

print(f"\n导入完成：成功 {success}/{len(rows)} 条")
if failures:
    print(f"失败 {len(failures)} 条：")
    for f in failures[:10]:
        print(f"  行{f[0]}: {f[1]} — {f[2]}")
    if len(failures) > 10:
        print(f"  ...还有 {len(failures) - 10} 条")
print(f"采购总额：¥{total_cost_sum:.2f}")
