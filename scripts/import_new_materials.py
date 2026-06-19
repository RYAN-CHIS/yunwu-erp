#!/usr/bin/env python3
"""
进货清单_新版.xlsx 导入脚本 V6（最终版）
修复三个问题:
  1. shape 直接存中文（圆珠/三通/挂饰），不做英文映射
  2. 颗数/条(beads_per_strand) 按串的正确填充
  3. 所有数值字段完整录入
"""

import openpyxl
import psycopg2
import re

EXCEL_PATH = "/Users/ryan/Desktop/进货清单_新版.xlsx"
SHEET_NAME = "01原料采购库"

def get_dsn():
    with open("/Users/ryan/yunwu-brand-os/.env") as f:
        for line in f:
            m = re.search(r'DATABASE_URL="(.+?)"', line)
            if m:
                return m.group(1)
    raise RuntimeError("DATABASE_URL not found")

def main():
    dsn = get_dsn()
    
    # 读取 Excel
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        code = row[0]
        if not code:
            continue
        rows.append(row)
    
    print(f"=== 读取 Excel: {len(rows)} 条有效数据 ===\n")
    wb.close()
    
    # 连接数据库
    conn = psycopg2.connect(dsn, connect_timeout=10)
    cur = conn.cursor()
    
    # 清空现有数据
    print("=== 清空现有数据 ===")
    for table in ['bom', 'inventory_transactions', 'purchase_records', 'raw_materials']:
        try:
            cur.execute(f'DELETE FROM "{table}"')
            print(f"  {table}: 已清空 ({cur.rowcount} 条)")
        except Exception as e:
            print(f"  {table}: 跳过 ({e})")
    
    conn.commit()
    print()
    
    print("=== 导入新材料数据 ===\n")
    
    mat_type_map = {
        '水晶': 'BEAD',
        '沉香': 'BEAD',
        '蜜蜡': 'BEAD',
        '青金石': 'BEAD',
        '南红': 'BEAD',
        '猛犸': 'BEAD',
        '老山檀': 'BEAD',
        '大漆珠': 'BEAD',
        '配件': 'METAL',
    }
    
    success = 0
    for idx, row in enumerate(rows):
        code          = str(row[0])   # A: 原料编码 (新格式)
        category      = row[1] or ''  # B: 品类
        name          = row[2] or ''  # C: 名称
        supplier      = row[3] or ''  # D: 供应商
        spec          = row[4]        # E: 规格mm
        shape         = str(row[5]) if row[5] else ''  # F: 形状（直接用中文!）
        remark        = row[6] or ''  # G: 备注
        price_mode    = (row[7] or '').strip()  # H: 计价方式
        qty           = row[8]        # I: 进货量(串数/个数/克重)
        unit_price    = row[9]        # J: 计价单价
        total_price   = row[11]       # L: 采购总价
        beads         = row[12]       # M: 每串颗数
        weight        = row[13]       # N: 每串克重
        unit_cost     = row[16]       # Q: 单颗成本
        
        if not name:
            continue
        
        material_type = mat_type_map.get(category, 'OTHER')
        
        # === 根据计价方式设置字段 ===
        if price_mode == '按克':
            inventory_unit = '克'
            purchase_unit = '克'
            conversion_rate = 1  # 采购=核算=克，无需换算
            beads_per_strand = None
            weight_per_strand = float(weight) if weight else None
            inv_qty = float(row[15]) if row[15] else 0  # P: 总克重 = 库存量
            uc = float(unit_cost) if unit_cost else None
            
        elif price_mode == '按个':
            inventory_unit = '个'
            purchase_unit = '个'
            conversion_rate = 1
            beads_per_strand = None
            weight_per_strand = None
            inv_qty = float(qty) if qty else 0
            # 按个：unitCost = 总价 / 个数
            if total_price and qty:
                uc = float(total_price) / float(qty)
            else:
                uc = None
                
        elif price_mode == '按串':
            inventory_unit = '颗'     # 核算单位 = 颗
            purchase_unit = '串'      # 采购单位 = 串
            bps = int(beads) if beads else None
            conversion_rate = float(bps) if bps else 1  # 换算率 = 每串颗数
            beads_per_strand = bps
            weight_per_strand = float(weight) if weight else None
            inv_qty = float(row[14]) if row[14] else 0  # O: 总颗数 = 库存量
            uc = float(unit_cost) if unit_cost else None
            
        else:
            print(f"  ⚠️ [{idx+1}] 未知计价方式 '{price_mode}': {code}")
            continue
        
        try:
            # 插入 raw_materials
            cur.execute("""
                INSERT INTO raw_materials (
                    code, name, category, "materialType", specification,
                    "inventoryUnit", "unitCost", remaining, status, shape,
                    "beads_per_strand", "weight_per_strand",
                    "default_purchase_unit", "default_conversion_rate",
                    supplier, remark, created_at, updated_at
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'ACTIVE',%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
                RETURNING id
            """, (
                code, name, category, material_type, str(spec) if spec else '',
                inventory_unit, uc, inv_qty, shape,
                beads_per_strand, weight_per_strand,
                purchase_unit, conversion_rate,
                supplier, remark
            ))
            
            mat_id = cur.fetchone()[0]
            
            # purchase_records
            p_qty = float(qty) if qty else 0
            p_unit_price = float(unit_price) if unit_price else 0
            p_total = float(total_price) if total_price else 0
            
            cur.execute("""
                INSERT INTO purchase_records (
                    "materialId", purchase_date, supplier,
                    "purchaseUnit", "conversionRate",
                    "purchaseQuantity", purchase_unit_price,
                    "purchasePrice", inventory_quantity, "unitCost",
                    remark, created_at
                ) VALUES (%s,NOW(),%s,%s,1,%s,%s,%s,%s,%s,%s,NOW())
            """, (
                mat_id, supplier, purchase_unit,
                p_qty, p_unit_price,
                p_total, inv_qty, uc,
                f"初始入库 {price_mode}"
            ))
            
            # inventory_transactions
            cur.execute("""
                INSERT INTO inventory_transactions (
                    "materialId", type, quantity,
                    "beforeQty", "afterQty",
                    remark, created_at
                ) VALUES (%s,'IN',%s,%s,%s,%s,NOW())
            """, (
                mat_id, inv_qty, 0, inv_qty,
                f"初始入库 {price_mode}"
            ))
            
            success += 1
            # 打印前5条 + 每10条 + 最后3条
            if success <= 5 or success % 15 == 0 or success >= len(rows) - 2:
                bps_str = f"{beads_per_strand}" if beads_per_strand else "-"
                wps_str = f"{weight_per_strand}g" if weight_per_strand else "-"
                print(f"  ✓ [{success:>2}] {code} | {name:<12} | {category:<4} | "
                      f"{shape:<4} | {price_mode:<2} | "
                      f"库存={inv_qty:.1f}{inventory_unit} | "
                      f"颗/条={bps_str} | 克/条={wps_str} | "
                      f"成本={'¥'+str(round(uc,2)) if uc else '-'}")
            
        except Exception as e:
            print(f"  ✗ {code} | {name}: {e}")
            conn.rollback()
            continue
    
    conn.commit()
    
    # ===== 验证 =====
    print(f"\n{'='*70}")
    print(f"导入完成: {success}/{len(rows)} 条")
    
    cur.execute("SELECT COUNT(*) FROM raw_materials")
    rm_count = cur.fetchone()[0]
    print(f"raw_materials: {rm_count}")
    cur.execute("SELECT COUNT(*) FROM purchase_records")
    pr_count = cur.fetchone()[0]
    print(f"purchase_records: {pr_count}")
    cur.execute("SELECT COUNT(*) FROM inventory_transactions")
    it_count = cur.fetchone()[0]
    print(f"inventory_transactions: {it_count}")
    
    # 形状分布（应全是中文）
    cur.execute("""
        SELECT shape, COUNT(*) FROM raw_materials 
        GROUP BY shape ORDER BY COUNT(*) DESC
    """)
    print("\n形状分布:")
    for s, c in cur.fetchall():
        print(f"  {s}: {c}条")
    
    # 计价方式分布
    cur.execute("""
        SELECT 
            CASE WHEN remark LIKE '%按克%' THEN '按克'
                 WHEN remark LIKE '%按个%' THEN '按个'
                 WHEN remark LIKE '%按串%' THEN '按串'
                 ELSE '其他' END as mode,
            COUNT(*)
        FROM raw_materials GROUP BY mode ORDER BY COUNT(*) DESC
    """)
    print("\n计价方式:")
    for mode, c in cur.fetchall():
        print(f"  {mode}: {c}条")
    
    # 检查关键字段完整性
    cur.execute("""
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN shape IS NOT NULL AND shape != '' THEN 1 ELSE 0 END) as has_shape,
            SUM(CASE WHEN \"unitCost\" IS NOT NULL THEN 1 ELSE 0 END) as has_cost,
            SUM(CASE WHEN \"beads_per_strand\" IS NOT NULL THEN 1 ELSE 0 END) as has_beads,
            SUM(CASE WHEN \"weight_per_strand\" IS NOT NULL THEN 1 ELSE 0 END) as has_weight,
            SUM(CASE WHEN \"default_conversion_rate\" > 1 THEN 1 ELSE 0 END) as rate_gt1,
            SUM(CASE WHEN remaining > 0 THEN 1 ELSE 0 END) as has_remaining
        FROM raw_materials
    """)
    r = cur.fetchone()
    print(f"\n数据完整性:")
    print(f"  总数: {r[0]}")
    print(f"  有形状: {r[1]}/{r[0]}")
    print(f"  有成本: {r[2]}/{r[0]}")
    print(f"  有颗/条: {r[3]}/{r[0]} (仅按串)")
    print(f"  有克/条: {r[4]}/{r[0]} (仅按克)")
    print(f"  换算率>1: {r[5]}/{r[0]} (按串才有)")
    print(f"  有库存: {r[6]}/{r[0]}")
    
    # 总金额
    cur.execute('SELECT COALESCE(SUM("purchasePrice"), 0)::numeric(12,2) FROM purchase_records')
    total = cur.fetchone()[0]
    print(f"\n总采购金额: ¥{total}")
    
    conn.close()
    print("\n✅ V6 导入完成！shape已改为中文")

if __name__ == "__main__":
    main()
